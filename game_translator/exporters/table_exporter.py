"""Table format exporters (Excel, CSV)"""

import csv
from pathlib import Path
from typing import Dict, Any, Optional

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from .base import BaseExporter


class ExcelExporter(BaseExporter):
    """Export to Excel format with formatting"""

    def __init__(self):
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

    def export(self, data: Dict[str, Any], output_path: Path,
               glossary: Optional[Dict[str, str]] = None):
        """Export to Excel with formatting and glossary sheet"""
        self.ensure_output_dir(output_path)

        wb = openpyxl.Workbook()

        # Main translations sheet
        ws = wb.active
        ws.title = "Translations"

        # Add headers
        headers = ["Key", "Context", "Source Text", "Translation", "Status", "Notes", "File"]
        self._add_headers(ws, headers)

        # Status colors
        status_colors = {
            "pending": "FFF2CC",      # Light yellow
            "translated": "D5E8D4",    # Light green
        }

        # Add data rows
        entries = data.get("entries", [])
        for row_idx, entry in enumerate(entries, 2):
            ws.cell(row=row_idx, column=1, value=entry.get("key", ""))
            ws.cell(row=row_idx, column=2, value=entry.get("context", ""))
            ws.cell(row=row_idx, column=3, value=entry.get("source", ""))
            ws.cell(row=row_idx, column=4, value=entry.get("translation", ""))

            # Status with color
            status = entry.get("status", "pending")
            status_cell = ws.cell(row=row_idx, column=5, value=status)
            if status in status_colors:
                status_cell.fill = PatternFill(
                    start_color=status_colors[status],
                    end_color=status_colors[status],
                    fill_type="solid"
                )

            ws.cell(row=row_idx, column=6, value=entry.get("notes", ""))
            ws.cell(row=row_idx, column=7, value=entry.get("file", ""))

        # Auto-adjust column widths
        self._adjust_column_widths(ws)

        # Add glossary sheet if provided
        if glossary:
            self._add_glossary_sheet(wb, glossary)

        # Add statistics sheet
        self._add_stats_sheet(wb, data)

        # Save workbook
        wb.save(output_path)
        print(f"Exported to Excel: {output_path}")

    def _add_headers(self, ws, headers):
        """Add formatted headers to worksheet"""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="366092",
                end_color="366092",
                fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

    def _adjust_column_widths(self, ws):
        """Auto-adjust column widths based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _add_glossary_sheet(self, wb, glossary):
        """Add glossary sheet with terms and translations"""
        ws = wb.create_sheet("Glossary")

        # Headers
        headers = ["Term", "Translation", "Notes"]
        self._add_headers(ws, headers)

        # Add glossary entries
        for row_idx, (term, translation) in enumerate(sorted(glossary.items()), 2):
            ws.cell(row=row_idx, column=1, value=term)
            ws.cell(row=row_idx, column=2, value=translation)
            ws.cell(row=row_idx, column=3, value="")  # Empty notes column for user

        self._adjust_column_widths(ws)

    def _add_stats_sheet(self, wb, data):
        """Add statistics sheet"""
        ws = wb.create_sheet("Statistics")

        stats = data.get("stats", {})
        info = [
            ("Project", data.get("project", "")),
            ("Source Language", data.get("source_lang", "")),
            ("Target Language", data.get("target_lang", "")),
            ("", ""),  # Empty row
            ("Total Entries", stats.get("total", 0)),
            ("Pending", stats.get("pending", 0)),
            ("Translated", stats.get("translated", 0)),
            ("Completion Rate", f"{stats.get('completion_rate', 0):.1f}%"),
        ]

        for row_idx, (label, value) in enumerate(info, 1):
            if label:  # Skip empty rows
                cell_label = ws.cell(row=row_idx, column=1, value=label)
                cell_label.font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=value)

        self._adjust_column_widths(ws)


class CsvExporter(BaseExporter):
    """Export to CSV format"""

    def export(self, data: Dict[str, Any], output_path: Path,
               glossary: Optional[Dict[str, str]] = None):
        """Export to CSV file"""
        self.ensure_output_dir(output_path)

        entries = data.get("entries", [])

        with open(output_path, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f)

            # Write headers
            writer.writerow(["Key", "Context", "Source", "Translation", "Status", "Notes"])

            # Write data
            for entry in entries:
                writer.writerow([
                    entry.get("key", ""),
                    entry.get("context", ""),
                    entry.get("source", ""),
                    entry.get("translation", ""),
                    entry.get("status", ""),
                    entry.get("notes", "")
                ])

        print(f"Exported to CSV: {output_path}")

        # Export glossary separately if provided
        if glossary:
            glossary_path = output_path.parent / f"{output_path.stem}_glossary.csv"
            self._export_glossary_csv(glossary, glossary_path)

    def _export_glossary_csv(self, glossary: Dict[str, str], output_path: Path):
        """Export glossary to separate CSV"""
        with open(output_path, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f)

            # Headers
            writer.writerow(["Term", "Translation"])

            # Data
            for term, translation in sorted(glossary.items()):
                writer.writerow([term, translation])

        print(f"Exported glossary to CSV: {output_path}")