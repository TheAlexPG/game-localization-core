"""Custom validation patterns management"""

import json
import csv
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import re


class CustomPatternsManager:
    """Manages custom validation patterns from various sources"""

    def __init__(self):
        self.patterns = {}

    def load_from_csv(self, csv_path: Path) -> Dict[str, Dict[str, str]]:
        """Load custom patterns from CSV file

        Expected CSV format:
        name,pattern,description,enabled
        my_marker,"\\[\\w+\\]","Square bracket markers",true
        special_id,"#\\d{4,6}","Special ID numbers",true
        """
        patterns = {}

        try:
            with open(csv_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)

                for row in reader:
                    name = row.get('name', '').strip()
                    pattern = row.get('pattern', '').strip()
                    description = row.get('description', '').strip()
                    enabled = row.get('enabled', 'true').lower() == 'true'

                    if name and pattern and enabled:
                        # Validate pattern
                        try:
                            re.compile(pattern)
                            patterns[name] = {
                                'pattern': pattern,
                                'description': description or f"Custom pattern: {name}"
                            }
                        except re.error as e:
                            print(f"Warning: Invalid pattern '{name}': {pattern} - {e}")

            self.patterns.update(patterns)
            return patterns

        except Exception as e:
            print(f"Error loading patterns from {csv_path}: {e}")
            return {}

    def load_from_excel(self, excel_path: Path, sheet_name: str = "ValidationPatterns") -> Dict[str, Dict[str, str]]:
        """Load custom patterns from Excel file

        Expected Excel sheet format:
        Column A: name
        Column B: pattern
        Column C: description
        Column D: enabled
        """
        patterns = {}

        try:
            # Try to import openpyxl
            try:
                from openpyxl import load_workbook
            except ImportError:
                print("Error: openpyxl not installed. Cannot read Excel files.")
                print("Install with: pip install openpyxl")
                return {}

            workbook = load_workbook(excel_path, read_only=True)

            if sheet_name not in workbook.sheetnames:
                print(f"Warning: Sheet '{sheet_name}' not found. Available sheets: {workbook.sheetnames}")
                return {}

            worksheet = workbook[sheet_name]

            # Skip header row (row 1)
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:  # Skip empty rows
                    continue

                name = str(row[0]).strip() if row[0] else ""
                pattern = str(row[1]).strip() if row[1] else ""
                description = str(row[2]).strip() if row[2] else ""
                enabled = str(row[3]).lower() == 'true' if row[3] else True

                if name and pattern and enabled:
                    # Validate pattern
                    try:
                        re.compile(pattern)
                        patterns[name] = {
                            'pattern': pattern,
                            'description': description or f"Custom pattern: {name}"
                        }
                    except re.error as e:
                        print(f"Warning: Invalid pattern '{name}': {pattern} - {e}")

            workbook.close()
            self.patterns.update(patterns)
            return patterns

        except Exception as e:
            print(f"Error loading patterns from {excel_path}: {e}")
            return {}

    def load_from_json(self, json_path: Path) -> Dict[str, Dict[str, str]]:
        """Load custom patterns from JSON file

        Expected JSON format:
        {
            "patterns": {
                "my_marker": {
                    "pattern": "\\\\[\\\\w+\\\\]",
                    "description": "Square bracket markers",
                    "enabled": true
                }
            }
        }
        """
        patterns = {}

        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            pattern_data = data.get('patterns', {})

            for name, info in pattern_data.items():
                if info.get('enabled', True):
                    pattern = info.get('pattern', '')
                    description = info.get('description', f"Custom pattern: {name}")

                    if pattern:
                        # Validate pattern
                        try:
                            re.compile(pattern)
                            patterns[name] = {
                                'pattern': pattern,
                                'description': description
                            }
                        except re.error as e:
                            print(f"Warning: Invalid pattern '{name}': {pattern} - {e}")

            self.patterns.update(patterns)
            return patterns

        except Exception as e:
            print(f"Error loading patterns from {json_path}: {e}")
            return {}

    def save_template_csv(self, csv_path: Path):
        """Create a template CSV file for custom patterns"""
        template_data = [
            {
                'name': 'square_brackets',
                'pattern': r'\[\w+\]',
                'description': 'Square bracket markers like [ACTION] [ITEM]',
                'enabled': 'true'
            },
            {
                'name': 'special_ids',
                'pattern': r'#\d{4,6}',
                'description': 'Special ID numbers like #1234 #123456',
                'enabled': 'true'
            },
            {
                'name': 'percentage_vars',
                'pattern': r'%\w+%',
                'description': 'Percentage variables like %PLAYER% %LEVEL%',
                'enabled': 'false'
            }
        ]

        try:
            with open(csv_path, 'w', encoding='utf-8', newline='') as f:
                writer = csv.DictWriter(f, fieldnames=['name', 'pattern', 'description', 'enabled'])
                writer.writeheader()
                writer.writerows(template_data)

            print(f"Template CSV created: {csv_path}")

        except Exception as e:
            print(f"Error creating template: {e}")

    def save_template_excel(self, excel_path: Path):
        """Create a template Excel file for custom patterns"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            print("Error: openpyxl not installed. Cannot create Excel files.")
            return

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "ValidationPatterns"

        # Headers
        headers = ['Name', 'Pattern', 'Description', 'Enabled']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Template data
        template_data = [
            ['square_brackets', r'\[\w+\]', 'Square bracket markers like [ACTION] [ITEM]', 'TRUE'],
            ['special_ids', r'#\d{4,6}', 'Special ID numbers like #1234 #123456', 'TRUE'],
            ['percentage_vars', r'%\w+%', 'Percentage variables like %PLAYER% %LEVEL%', 'FALSE'],
            ['unity_refs', r'{{[^}]+}}', 'Unity text mesh references like {{ref}}', 'FALSE'],
            ['color_codes', r'\[color=#[0-9a-fA-F]{6}\]', 'Color codes like [color=#FF0000]', 'FALSE']
        ]

        for row, data in enumerate(template_data, 2):
            for col, value in enumerate(data, 1):
                worksheet.cell(row=row, column=col, value=value)

        # Adjust column widths
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 30
        worksheet.column_dimensions['C'].width = 50
        worksheet.column_dimensions['D'].width = 10

        try:
            workbook.save(excel_path)
            print(f"Template Excel created: {excel_path}")
        except Exception as e:
            print(f"Error saving Excel template: {e}")

    def get_patterns_for_validator(self) -> Dict[str, str]:
        """Get patterns in format suitable for TranslationValidator"""
        return {name: info['pattern'] for name, info in self.patterns.items()}

    def list_patterns(self) -> List[Tuple[str, str, str]]:
        """List all loaded patterns"""
        return [(name, info['pattern'], info['description'])
                for name, info in self.patterns.items()]